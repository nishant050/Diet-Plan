"""
Diet Plan Dashboard — Main Application
FastAPI + Jinja2 + HTMX + MongoDB (Motor)
"""
import os
import io
import csv
import json
from collections import defaultdict
from datetime import datetime, date, timedelta
from contextlib import asynccontextmanager

from fastapi import FastAPI, Request, Form, UploadFile, File, Response, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from passlib.hash import bcrypt
from bson import ObjectId
from pymongo import ASCENDING, DESCENDING

from database import init_db, get_db
from ai_service import query_ai, get_all_models


# ──────────────────────── App Setup ────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    yield

app = FastAPI(title="Diet Plan Dashboard", lifespan=lifespan)

# Static files & templates
os.makedirs("static", exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
templates.env.globals["timedelta"] = timedelta
templates.env.globals["date"] = date
templates.env.globals["str"] = str


# ──────────────────────── Helpers ────────────────────────
def parse_user_agent(ua: str) -> dict:
    """Basic user agent parsing"""
    os_info = "Unknown"
    browser = "Unknown"
    device = "Unknown"

    ua_lower = ua.lower()

    # OS detection
    if "windows" in ua_lower:
        os_info = "Windows"
    elif "mac os" in ua_lower or "macintosh" in ua_lower:
        os_info = "macOS"
    elif "linux" in ua_lower:
        os_info = "Linux"
    elif "android" in ua_lower:
        os_info = "Android"
    elif "iphone" in ua_lower or "ipad" in ua_lower:
        os_info = "iOS"

    # Browser detection
    if "edg/" in ua_lower:
        browser = "Edge"
    elif "chrome" in ua_lower and "safari" in ua_lower:
        browser = "Chrome"
    elif "firefox" in ua_lower:
        browser = "Firefox"
    elif "safari" in ua_lower:
        browser = "Safari"
    elif "opera" in ua_lower or "opr/" in ua_lower:
        browser = "Opera"

    # Device type
    if "mobile" in ua_lower or "android" in ua_lower or "iphone" in ua_lower:
        device = "Mobile"
    elif "tablet" in ua_lower or "ipad" in ua_lower:
        device = "Tablet"
    else:
        device = "Desktop"

    return {"os": os_info, "browser": browser, "device": device}


async def log_activity(profile_id, action, details, request: Request):
    db = await get_db()
    ua = request.headers.get("user-agent", "")
    ua_info = parse_user_agent(ua)
    fp = request.cookies.get("device_fp", "")
    ip = request.client.host if request.client else ""

    await db.activity_logs.insert_one({
        "profile_id": str(profile_id) if profile_id else None,
        "action": action,
        "details": details,
        "device_fingerprint": fp,
        "user_agent": ua,
        "os_info": ua_info["os"],
        "browser_info": ua_info["browser"],
        "device_type": ua_info["device"],
        "ip_address": ip,
        "created_at": datetime.utcnow()
    })


def get_admin_session(request: Request):
    return request.cookies.get("admin_session")


MEAL_TYPE_ORDER = ['breakfast', 'morning_snack', 'lunch', 'afternoon_snack', 'dinner', 'evening_snack']
MEAL_TYPE_LABELS = {
    'breakfast': '🌅 Breakfast',
    'morning_snack': '🍎 Morning Snack',
    'lunch': '☀️ Lunch',
    'afternoon_snack': '🫐 Afternoon Snack',
    'dinner': '🌙 Dinner',
    'evening_snack': '🌜 Evening Snack'
}


# ──────────────────────── User Routes ────────────────────────

def parse_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def parse_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def sort_meal_type(meal_type: str) -> int:
    if meal_type in MEAL_TYPE_ORDER:
        return MEAL_TYPE_ORDER.index(meal_type)
    return len(MEAL_TYPE_ORDER)


def serialize_profile(profile: dict) -> dict:
    return {
        "id": str(profile["_id"]),
        "name": profile.get("name", "Unnamed"),
        "created_at": profile.get("created_at"),
    }


def normalize_meal_doc(meal: dict, profile_map: dict) -> dict:
    profile_id = str(meal.get("profile_id")) if meal.get("profile_id") else "0"
    calories = parse_int(meal.get("calories", 0))
    protein_g = parse_float(meal.get("protein_g", 0))
    carbs_g = parse_float(meal.get("carbs_g", 0))
    fat_g = parse_float(meal.get("fat_g", 0))
    fiber_g = parse_float(meal.get("fiber_g", 0))
    is_personal = profile_id != "0"

    return {
        "id": str(meal["_id"]),
        "plan_date": meal.get("plan_date", ""),
        "meal_type": meal.get("meal_type", ""),
        "meal_type_label": MEAL_TYPE_LABELS.get(
            meal.get("meal_type", ""),
            str(meal.get("meal_type", "")).replace("_", " ").title(),
        ),
        "dish_name": meal.get("dish_name", "").strip(),
        "description": meal.get("description", "").strip(),
        "calories": calories,
        "protein_g": protein_g,
        "carbs_g": carbs_g,
        "fat_g": fat_g,
        "fiber_g": fiber_g,
        "profile_id": profile_id,
        "assigned_to": profile_map.get(profile_id, "Everyone") if is_personal else "Everyone",
        "assignment_scope": "Assigned" if is_personal else "Everyone",
        "is_personal": is_personal,
    }


def build_day_sections(meals: list[dict], view_all: int, today: date, week_start: date | None = None) -> list[dict]:
    buckets: dict[str, list[dict]] = defaultdict(list)
    for meal in meals:
        buckets[meal["plan_date"]].append(meal)

    day_keys = sorted(buckets.keys())
    if not view_all and week_start is not None:
        day_keys = [(week_start + timedelta(days=offset)).isoformat() for offset in range(7)]

    sections = []
    for day_key in day_keys:
        day_date = date.fromisoformat(day_key)
        day_meals = sorted(buckets.get(day_key, []), key=lambda item: sort_meal_type(item["meal_type"]))
        sections.append({
            "iso": day_key,
            "date": day_date,
            "label": day_date.strftime("%a %d %b"),
            "long_label": day_date.strftime("%A, %d %b %Y"),
            "is_today": day_date == today,
            "meal_count": len(day_meals),
            "total_calories": sum(item["calories"] for item in day_meals),
            "meals": day_meals,
        })

    return sections


async def load_meal_templates(db, limit: int = 8) -> list[dict]:
    pipeline = [
        {"$sort": {"created_at": DESCENDING}},
        {
            "$group": {
                "_id": {
                    "dish_name": "$dish_name",
                    "meal_type": "$meal_type",
                    "profile_id": "$profile_id",
                },
                "dish_name": {"$first": "$dish_name"},
                "meal_type": {"$first": "$meal_type"},
                "description": {"$first": "$description"},
                "calories": {"$first": "$calories"},
                "protein_g": {"$first": "$protein_g"},
                "carbs_g": {"$first": "$carbs_g"},
                "fat_g": {"$first": "$fat_g"},
                "fiber_g": {"$first": "$fiber_g"},
                "profile_id": {"$first": "$profile_id"},
                "uses": {"$sum": 1},
            }
        },
        {"$sort": {"uses": DESCENDING, "dish_name": ASCENDING}},
        {"$limit": limit},
    ]

    docs = await db.meal_plans.aggregate(pipeline).to_list(length=limit)
    templates_list = []
    for doc in docs:
        profile_id = str(doc.get("profile_id")) if doc.get("profile_id") else "0"
        templates_list.append({
            "dish_name": doc.get("dish_name", ""),
            "meal_type": doc.get("meal_type", "lunch"),
            "meal_type_label": MEAL_TYPE_LABELS.get(
                doc.get("meal_type", "lunch"),
                str(doc.get("meal_type", "lunch")).replace("_", " ").title(),
            ),
            "description": doc.get("description", "") or "",
            "calories": parse_int(doc.get("calories", 0)),
            "protein_g": parse_float(doc.get("protein_g", 0)),
            "carbs_g": parse_float(doc.get("carbs_g", 0)),
            "fat_g": parse_float(doc.get("fat_g", 0)),
            "fiber_g": parse_float(doc.get("fiber_g", 0)),
            "profile_id": profile_id,
            "uses": doc.get("uses", 1),
        })

    return templates_list


async def build_admin_meals_context(
    request: Request,
    week_offset: int = 0,
    view_all: int = 0,
    meal_type_filter: str = "all",
    assignee_filter: str = "all",
    q: str = "",
    flash_message: dict | None = None,
):
    db = await get_db()
    today = date.today()

    profile_docs = await db.profiles.find().sort("name", ASCENDING).to_list(length=None)
    profiles = [serialize_profile(profile) for profile in profile_docs]
    profile_map = {profile["id"]: profile["name"] for profile in profiles}

    week_start = None
    week_end = None
    scope_query = {}
    if not view_all:
        week_start = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
        week_end = week_start + timedelta(days=6)
        scope_query["plan_date"] = {"$gte": week_start.isoformat(), "$lte": week_end.isoformat()}

    query = dict(scope_query)
    if meal_type_filter != "all":
        query["meal_type"] = meal_type_filter

    if assignee_filter == "everyone":
        query["profile_id"] = None
    elif assignee_filter != "all":
        query["profile_id"] = assignee_filter

    if q.strip():
        query["$or"] = [
            {"dish_name": {"$regex": q.strip(), "$options": "i"}},
            {"description": {"$regex": q.strip(), "$options": "i"}},
        ]

    scope_total = await db.meal_plans.count_documents(scope_query)
    visible_total = await db.meal_plans.count_documents(query)
    raw_meals = await db.meal_plans.find(query).sort([("plan_date", ASCENDING), ("meal_type", ASCENDING)]).to_list(length=None)
    meals = [normalize_meal_doc(meal, profile_map) for meal in raw_meals]
    day_sections = build_day_sections(meals, view_all, today, week_start)

    summary = {
        "visible_total": visible_total,
        "scope_total": scope_total,
        "day_count": len([section for section in day_sections if section["meal_count"] > 0]) if view_all else len(day_sections),
        "everyone_total": sum(1 for meal in meals if not meal["is_personal"]),
        "assigned_total": sum(1 for meal in meals if meal["is_personal"]),
        "total_calories": sum(meal["calories"] for meal in meals),
    }

    return {
        "request": request,
        "today": today,
        "meals": meals,
        "day_sections": day_sections,
        "profiles": profiles,
        "templates_list": await load_meal_templates(db),
        "summary": summary,
        "week_offset": week_offset,
        "week_start": week_start,
        "week_end": week_end,
        "view_all": view_all,
        "meal_type_labels": MEAL_TYPE_LABELS,
        "meal_type_order": MEAL_TYPE_ORDER,
        "filters": {
            "meal_type_filter": meal_type_filter,
            "assignee_filter": assignee_filter,
            "q": q.strip(),
        },
        "flash_message": flash_message,
    }


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Landing page — checks for known profile via cookie"""
    profile_id = request.cookies.get("profile_id")
    if profile_id:
        return RedirectResponse(url="/dashboard", status_code=302)
    return templates.TemplateResponse("welcome.html", {"request": request})


@app.post("/profile/new", response_class=HTMLResponse)
async def create_profile(request: Request, name: str = Form(...)):
    db = await get_db()
    
    fp = request.cookies.get("device_fp", "")
    res = await db.profiles.insert_one({
        "name": name.strip(),
        "device_fingerprint": fp,
        "created_at": datetime.utcnow()
    })
    profile_id = str(res.inserted_id)

    # Link device to profile
    if fp:
        await db.device_profile_map.insert_one({
            "device_fingerprint": fp,
            "profile_id": profile_id,
            "created_at": datetime.utcnow()
        })

    await log_activity(profile_id, "profile_created", f"New profile: {name}", request)

    response = RedirectResponse(url="/dashboard", status_code=302)
    response.set_cookie("profile_id", profile_id, max_age=365*24*3600, samesite="lax")
    return response


@app.get("/profiles/list", response_class=HTMLResponse)
async def list_profiles(request: Request):
    db = await get_db()
    profiles = await db.profiles.find().sort("name", ASCENDING).to_list(length=None)
    
    # Cast _id to string for the template
    for p in profiles:
        p["id"] = str(p["_id"])
        
    return templates.TemplateResponse("profiles_list.html", {"request": request, "profiles": profiles})


@app.post("/profile/select/{profile_id}")
async def select_profile(request: Request, profile_id: str):
    db = await get_db()
    try:
        profile = await db.profiles.find_one({"_id": ObjectId(profile_id)})
    except Exception:
        raise HTTPException(404, "Invalid Profile ID format")

    if not profile:
        raise HTTPException(404, "Profile not found")

    # Link device
    fp = request.cookies.get("device_fp", "")
    if fp:
        # Check if exists
        existing = await db.device_profile_map.find_one({"device_fingerprint": fp, "profile_id": profile_id})
        if not existing:
            await db.device_profile_map.insert_one({
                "device_fingerprint": fp,
                "profile_id": profile_id,
                "created_at": datetime.utcnow()
            })

    await log_activity(profile_id, "profile_selected", f"Selected profile: {profile['name']}", request)

    response = RedirectResponse(url="/dashboard", status_code=302)
    response.set_cookie("profile_id", profile_id, max_age=365*24*3600, samesite="lax")
    return response


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, view_date: str = None):
    profile_id = request.cookies.get("profile_id")
    if not profile_id:
        return RedirectResponse(url="/", status_code=302)

    today = date.today()
    if view_date:
        try:
            selected_date = date.fromisoformat(view_date)
        except ValueError:
            selected_date = today
    else:
        selected_date = today

    db = await get_db()
    
    # Get profile
    try:
        profile = await db.profiles.find_one({"_id": ObjectId(profile_id)})
    except Exception:
        response = RedirectResponse(url="/", status_code=302)
        response.delete_cookie("profile_id")
        return response
        
    if not profile:
        response = RedirectResponse(url="/", status_code=302)
        response.delete_cookie("profile_id")
        return response

    # If no specific date requested, check if today has meals; if not, find nearest date
    if not view_date:
        count = await db.meal_plans.count_documents({
            "plan_date": today.isoformat(),
            "$or": [{"profile_id": None}, {"profile_id": profile_id}]
        })
        
        if count == 0:
            # Find closest date - sort of complex query in Mongo, let's just find the next upcoming one or failing that, most recent past one
            upcoming = await db.meal_plans.find({
                "plan_date": {"$gt": today.isoformat()},
                "$or": [{"profile_id": None}, {"profile_id": profile_id}]
            }).sort("plan_date", ASCENDING).limit(1).to_list(1)
            
            if upcoming:
                selected_date = date.fromisoformat(upcoming[0]["plan_date"])
            else:
                past = await db.meal_plans.find({
                    "plan_date": {"$lt": today.isoformat()},
                    "$or": [{"profile_id": None}, {"profile_id": profile_id}]
                }).sort("plan_date", DESCENDING).limit(1).to_list(1)
                
                if past:
                    selected_date = date.fromisoformat(past[0]["plan_date"])

    # Get meals for the day (user's meals + global meals)
    meals = await db.meal_plans.find({
        "plan_date": selected_date.isoformat(),
        "$or": [{"profile_id": None}, {"profile_id": profile_id}]
    }).to_list(length=None)

    # Sort meals conceptually based on order array since we lost SQL ORDER BY CASE
    meals.sort(key=lambda m: MEAL_TYPE_ORDER.index(m["meal_type"]) if m["meal_type"] in MEAL_TYPE_ORDER else 99)

    # Get checked meals
    checks = await db.meal_checks.find({
        "profile_id": profile_id,
        "is_prepared": 1
    }).to_list(length=None)
    checked_ids = {str(c["meal_plan_id"]) for c in checks}

    # Organize meals by type
    organized = {}
    for mt in MEAL_TYPE_ORDER:
        organized[mt] = {
            "label": MEAL_TYPE_LABELS[mt],
            "meals": []
        }

    for meal in meals:
        mt = meal["meal_type"]
        meal_id_str = str(meal["_id"])
        if mt in organized:
            organized[mt]["meals"].append({
                "id": meal_id_str,
                "dish_name": meal["dish_name"],
                "description": meal.get("description", ""),
                "calories": meal.get("calories", 0),
                "protein_g": meal.get("protein_g", 0),
                "carbs_g": meal.get("carbs_g", 0),
                "fat_g": meal.get("fat_g", 0),
                "fiber_g": meal.get("fiber_g", 0),
                "is_checked": meal_id_str in checked_ids
            })

    # Compute total calories for the day
    total_cal = sum(m.get("calories", 0) for m in meals)
    checked_cal = sum(m.get("calories", 0) for m in meals if str(m["_id"]) in checked_ids)

    await log_activity(profile_id, "page_view", f"Dashboard for {selected_date.isoformat()}", request)

    # Navigation dates
    prev_date = (selected_date - timedelta(days=1)).isoformat()
    next_date = (selected_date + timedelta(days=1)).isoformat()
    is_today = selected_date == today

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "profile": profile,
        "meals": organized,
        "selected_date": selected_date,
        "today": today,
        "prev_date": prev_date,
        "next_date": next_date,
        "is_today": is_today,
        "total_calories": total_cal,
        "checked_calories": checked_cal,
        "meal_type_order": MEAL_TYPE_ORDER,
    })


@app.post("/meal/toggle/{meal_plan_id}", response_class=HTMLResponse)
async def toggle_meal(request: Request, meal_plan_id: str):
    profile_id = request.cookies.get("profile_id")
    if not profile_id:
        return HTMLResponse("<span>Error</span>", status_code=401)

    db = await get_db()
    
    existing = await db.meal_checks.find_one({
        "profile_id": profile_id,
        "meal_plan_id": meal_plan_id
    })

    if existing:
        new_val = 0 if existing.get("is_prepared") else 1
        await db.meal_checks.update_one(
            {"_id": existing["_id"]},
            {"$set": {"is_prepared": new_val, "checked_at": datetime.utcnow()}}
        )
    else:
        new_val = 1
        await db.meal_checks.insert_one({
            "profile_id": profile_id,
            "meal_plan_id": meal_plan_id,
            "is_prepared": 1,
            "checked_at": datetime.utcnow()
        })

    # Get dish name for logging
    try:
        mp = await db.meal_plans.find_one({"_id": ObjectId(meal_plan_id)})
        dish = mp["dish_name"] if mp else "Unknown"
    except Exception:
        dish = "Unknown"

    action = "meal_prepared" if new_val else "meal_unprepared"
    await log_activity(profile_id, action, dish, request)

    icon = "✅" if new_val else "⬜"
    return HTMLResponse(f"""<span class="check-icon">{icon}</span>""")


@app.get("/dish/info/{meal_plan_id}", response_class=HTMLResponse)
async def dish_info(request: Request, meal_plan_id: str):
    profile_id = request.cookies.get("profile_id")

    db = await get_db()
    try:
        meal = await db.meal_plans.find_one({"_id": ObjectId(meal_plan_id)})
    except Exception:
        return HTMLResponse("<p>Invalid Dish ID format</p>")
        
    if not meal:
        return HTMLResponse("<p>Dish not found</p>")

    if profile_id:
        await log_activity(profile_id, "dish_viewed", meal["dish_name"], request)

    # Query AI
    result = await query_ai(meal["dish_name"])
    desc = meal.get("description", "")
    return HTMLResponse(f"""
        <div class="dish-detail-content">
            <div class="dish-detail-header">
                <h2>{meal["dish_name"]}</h2>
                <p class="dish-description">{desc}</p>
            </div>
            <div class="ai-content">
                {result}
            </div>
        </div>
    """)


@app.get("/history", response_class=HTMLResponse)
async def history(request: Request, week_offset: int = 0):
    profile_id = request.cookies.get("profile_id")
    if not profile_id:
        return RedirectResponse(url="/", status_code=302)

    today = date.today()
    # Calculate week start (Monday)
    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    end_of_week = start_of_week + timedelta(days=6)

    db = await get_db()
    
    try:
        profile = await db.profiles.find_one({"_id": ObjectId(profile_id)})
    except Exception:
        return RedirectResponse(url="/", status_code=302)

    # Get all meals for the week (user's meals + global meals)
    meals = await db.meal_plans.find({
        "plan_date": {"$gte": start_of_week.isoformat(), "$lte": end_of_week.isoformat()},
        "$or": [{"profile_id": None}, {"profile_id": profile_id}]
    }).sort([("plan_date", ASCENDING), ("meal_type", ASCENDING)]).to_list(length=None)

    # Get checked meals
    checks = await db.meal_checks.find({
        "profile_id": profile_id,
        "is_prepared": 1
    }).to_list(length=None)
    checked_ids = {str(row["meal_plan_id"]) for row in checks}

    # Organize by day
    days = {}
    for i in range(7):
        d = start_of_week + timedelta(days=i)
        days[d.isoformat()] = {
            "date": d,
            "day_name": d.strftime("%A"),
            "meals": [],
            "total": 0,
            "prepared": 0,
        }

    for meal in meals:
        d_key = meal["plan_date"]
        if d_key in days:
            is_checked = str(meal["_id"]) in checked_ids
            days[d_key]["meals"].append({
                "dish_name": meal["dish_name"],
                "meal_type": MEAL_TYPE_LABELS.get(meal["meal_type"], meal["meal_type"]),
                "calories": meal.get("calories", 0),
                "is_checked": is_checked,
            })
            days[d_key]["total"] += 1
            if is_checked:
                days[d_key]["prepared"] += 1

    await log_activity(profile_id, "history_view", f"Week of {start_of_week.isoformat()}", request)

    return templates.TemplateResponse("history.html", {
        "request": request,
        "profile": profile,
        "days": days,
        "start_of_week": start_of_week,
        "end_of_week": end_of_week,
        "prev_offset": week_offset - 1,
        "next_offset": week_offset + 1,
        "current_offset": week_offset,
        "today": today,
    })


@app.get("/profile/switch")
async def switch_profile(request: Request):
    response = RedirectResponse(url="/", status_code=302)
    response.delete_cookie("profile_id")
    return response


# ──────────────────────── Admin Routes ────────────────────────

@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    if get_admin_session(request):
        return RedirectResponse(url="/admin", status_code=302)
    return templates.TemplateResponse("admin_login.html", {"request": request})


@app.post("/admin/login")
async def admin_login(request: Request, username: str = Form(...), password: str = Form(...)):
    db = await get_db()
    
    admin = await db.admin.find_one({"username": username})
    if not admin or not bcrypt.verify(password, admin["password_hash"]):
        return templates.TemplateResponse("admin_login.html", {
            "request": request,
            "error": "Invalid credentials"
        })

    response = RedirectResponse(url="/admin", status_code=302)
    response.set_cookie("admin_session", "authenticated", max_age=8*3600, httponly=True, samesite="lax")
    return response


@app.get("/admin/logout")
async def admin_logout():
    response = RedirectResponse(url="/admin/login", status_code=302)
    response.delete_cookie("admin_session")
    return response


@app.get("/admin", response_class=HTMLResponse)
async def admin_dashboard(request: Request):
    if not get_admin_session(request):
        return RedirectResponse(url="/admin/login", status_code=302)

    db = await get_db()
    
    # Stats
    total_users = await db.profiles.count_documents({})
    total_meals = await db.meal_plans.count_documents({})
    total_days = len(await db.meal_plans.distinct("plan_date"))

    # Recent activity - Needs a manual lookup join equivalent since MongoDB doesn't easily JOIN outside of aggregations
    recent_acts = await db.activity_logs.find().sort("created_at", DESCENDING).limit(50).to_list(50)
    
    # Pre-fetch profiles for mapping
    all_profiles = await db.profiles.find().sort("name", ASCENDING).to_list(length=None)
    profile_map = {str(p["_id"]): p["name"] for p in all_profiles}
    
    # Attach profile_name manually
    activities = []
    for act in recent_acts:
        act["profile_name"] = profile_map.get(str(act.get("profile_id")), "Unknown Profile")
        activities.append(act)
        
    for p in all_profiles:
        p["id"] = str(p["_id"])

    # AI Models
    models = await get_all_models()
    for model in models:
        model["id"] = str(model["_id"])

    return templates.TemplateResponse("admin_dashboard.html", {
        "request": request,
        "total_users": total_users,
        "total_meals": total_meals,
        "total_days": total_days,
        "activities": activities,
        "profiles": all_profiles,
        "models": models,
    })


@app.get("/admin/template/download")
async def download_template(request: Request):
    if not get_admin_session(request):
        return RedirectResponse(url="/admin/login", status_code=302)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "plan_date", "meal_type", "dish_name", "description",
        "calories", "protein_g", "carbs_g", "fat_g", "fiber_g"
    ])
    # Sample rows
    writer.writerow([
        "2026-02-22", "breakfast", "Oatmeal with Berries",
        "Warm oatmeal topped with fresh berries and honey",
        "350", "12", "55", "8", "6"
    ])
    writer.writerow([
        "2026-02-22", "lunch", "Grilled Chicken Salad",
        "Mixed greens with grilled chicken breast",
        "420", "35", "15", "22", "4"
    ])
    writer.writerow([
        "2026-02-22", "dinner", "Salmon with Vegetables",
        "Baked salmon fillet with roasted seasonal vegetables",
        "480", "38", "20", "28", "5"
    ])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=meal_plan_template.csv"}
    )


@app.post("/admin/upload", response_class=HTMLResponse)
async def upload_meals(request: Request, file: UploadFile = File(...)):
    if not get_admin_session(request):
        return HTMLResponse("<p class='error'>Unauthorized</p>", status_code=401)

    content = await file.read()
    filename = file.filename.lower()

    rows = []
    errors = []

    try:
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for i, row in enumerate(reader, 2):
                rows.append(row)
        elif filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [str(cell.value).strip().lower() for cell in ws[1]]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if all(v is None for v in row):
                    continue
                rows.append(dict(zip(headers, row)))
        else:
            return HTMLResponse("""
                <div class="upload-result error">
                    <span class="icon">❌</span>
                    <p>Unsupported file format. Use CSV or XLSX.</p>
                </div>
            """)
    except Exception as e:
        return HTMLResponse(f"""
            <div class="upload-result error">
                <span class="icon">❌</span>
                <p>Error reading file: {str(e)}</p>
            </div>
        """)

    valid_types = {'breakfast', 'morning_snack', 'lunch', 'afternoon_snack', 'dinner', 'evening_snack'}
    
    db = await get_db()
    docs_to_insert = []
    
    for i, row in enumerate(rows, 2):
        try:
            plan_date = str(row.get("plan_date", "")).strip()
            meal_type = str(row.get("meal_type", "")).strip().lower()
            dish_name = str(row.get("dish_name", "")).strip()

            if not plan_date or not meal_type or not dish_name:
                errors.append(f"Row {i}: Missing required fields")
                continue

            if meal_type not in valid_types:
                errors.append(f"Row {i}: Invalid meal_type '{meal_type}'")
                continue

            # Validate date
            date.fromisoformat(plan_date)

            docs_to_insert.append({
                "plan_date": plan_date,
                "meal_type": meal_type,
                "dish_name": dish_name,
                "description": str(row.get("description", "")).strip(),
                "calories": int(float(row.get("calories", 0) or 0)),
                "protein_g": float(row.get("protein_g", 0) or 0),
                "carbs_g": float(row.get("carbs_g", 0) or 0),
                "fat_g": float(row.get("fat_g", 0) or 0),
                "fiber_g": float(row.get("fiber_g", 0) or 0),
                "profile_id": None,
                "created_at": datetime.utcnow()
            })
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    inserted = 0
    if docs_to_insert:
        res = await db.meal_plans.insert_many(docs_to_insert)
        inserted = len(res.inserted_ids)

    error_html = ""
    if errors:
        error_html = "<ul class='error-list'>" + "".join(f"<li>{e}</li>" for e in errors[:10]) + "</ul>"
        if len(errors) > 10:
            error_html += f"<p>...and {len(errors)-10} more errors</p>"

    return HTMLResponse(f"""
        <div class="upload-result {'success' if inserted > 0 else 'warning'}">
            <span class="icon">{'✅' if inserted > 0 else '⚠️'}</span>
            <p><strong>{inserted}</strong> meals uploaded successfully.</p>
            {f'<p class="error-count">{len(errors)} errors found:</p>{error_html}' if errors else ''}
        </div>
    """)


@app.get("/admin/meals", response_class=HTMLResponse)
async def admin_meals(
    request: Request,
    week_offset: int = 0,
    view_all: int = 0,
    meal_type_filter: str = "all",
    assignee_filter: str = "all",
    q: str = "",
):
    if not get_admin_session(request):
        return RedirectResponse(url="/admin/login", status_code=302)

    context = await build_admin_meals_context(
        request,
        week_offset=week_offset,
        view_all=view_all,
        meal_type_filter=meal_type_filter,
        assignee_filter=assignee_filter,
        q=q,
    )
    return templates.TemplateResponse("admin_meals.html", context)


@app.get("/admin/meals/board", response_class=HTMLResponse)
async def admin_meals_board(
    request: Request,
    week_offset: int = 0,
    view_all: int = 0,
    meal_type_filter: str = "all",
    assignee_filter: str = "all",
    q: str = "",
):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    context = await build_admin_meals_context(
        request,
        week_offset=week_offset,
        view_all=view_all,
        meal_type_filter=meal_type_filter,
        assignee_filter=assignee_filter,
        q=q,
    )
    return templates.TemplateResponse("admin_meals_board.html", context)


@app.post("/admin/meal/save", response_class=HTMLResponse)
async def save_meal(
    request: Request,
    meal_id: str = Form(""),
    plan_date: str = Form(...),
    meal_type: str = Form(...),
    dish_name: str = Form(...),
    description: str = Form(""),
    calories: float = Form(0),
    protein_g: float = Form(0),
    carbs_g: float = Form(0),
    fat_g: float = Form(0),
    fiber_g: float = Form(0),
    profile_id: str = Form("0"),
    week_offset: int = Form(0),
    view_all: int = Form(0),
    meal_type_filter: str = Form("all"),
    assignee_filter: str = Form("all"),
    q: str = Form(""),
):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    payload = {
        "plan_date": plan_date,
        "meal_type": meal_type,
        "dish_name": dish_name.strip(),
        "description": description.strip(),
        "calories": parse_int(calories),
        "protein_g": parse_float(protein_g),
        "carbs_g": parse_float(carbs_g),
        "fat_g": parse_float(fat_g),
        "fiber_g": parse_float(fiber_g),
        "profile_id": profile_id if profile_id and profile_id != "0" else None,
    }

    try:
        date.fromisoformat(plan_date)
    except ValueError:
        context = await build_admin_meals_context(
            request,
            week_offset=week_offset,
            view_all=view_all,
            meal_type_filter=meal_type_filter,
            assignee_filter=assignee_filter,
            q=q,
            flash_message={"kind": "error", "text": "Enter a valid date in YYYY-MM-DD format."},
        )
        return templates.TemplateResponse("admin_meals_board.html", context, status_code=422)

    if not payload["dish_name"]:
        context = await build_admin_meals_context(
            request,
            week_offset=week_offset,
            view_all=view_all,
            meal_type_filter=meal_type_filter,
            assignee_filter=assignee_filter,
            q=q,
            flash_message={"kind": "error", "text": "Dish name is required."},
        )
        return templates.TemplateResponse("admin_meals_board.html", context, status_code=422)

    if meal_id:
        try:
            await db.meal_plans.update_one(
                {"_id": ObjectId(meal_id)},
                {"$set": payload},
            )
            flash_text = f"Updated {payload['dish_name']}."
        except Exception:
            context = await build_admin_meals_context(
                request,
                week_offset=week_offset,
                view_all=view_all,
                meal_type_filter=meal_type_filter,
                assignee_filter=assignee_filter,
                q=q,
                flash_message={"kind": "error", "text": "Meal update failed."},
            )
            return templates.TemplateResponse("admin_meals_board.html", context, status_code=422)
    else:
        await db.meal_plans.insert_one({
            **payload,
            "created_at": datetime.utcnow(),
        })
        flash_text = f"Added {payload['dish_name']}."

    context = await build_admin_meals_context(
        request,
        week_offset=week_offset,
        view_all=view_all,
        meal_type_filter=meal_type_filter,
        assignee_filter=assignee_filter,
        q=q,
        flash_message={"kind": "success", "text": flash_text},
    )
    return templates.TemplateResponse("admin_meals_board.html", context)


@app.post("/admin/meal/{meal_id}/delete", response_class=HTMLResponse)
async def delete_meal_board(
    request: Request,
    meal_id: str,
    week_offset: int = Form(0),
    view_all: int = Form(0),
    meal_type_filter: str = Form("all"),
    assignee_filter: str = Form("all"),
    q: str = Form(""),
):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    await db.meal_checks.delete_many({"meal_plan_id": meal_id})
    try:
        await db.meal_plans.delete_one({"_id": ObjectId(meal_id)})
        flash_message = {"kind": "success", "text": "Meal deleted."}
    except Exception:
        flash_message = {"kind": "error", "text": "Meal delete failed."}

    context = await build_admin_meals_context(
        request,
        week_offset=week_offset,
        view_all=view_all,
        meal_type_filter=meal_type_filter,
        assignee_filter=assignee_filter,
        q=q,
        flash_message=flash_message,
    )
    return templates.TemplateResponse("admin_meals_board.html", context)


@app.post("/admin/meals/clear-range", response_class=HTMLResponse)
async def clear_meal_range(
    request: Request,
    week_start: str = Form(...),
    week_end: str = Form(...),
    week_offset: int = Form(0),
    view_all: int = Form(0),
    meal_type_filter: str = Form("all"),
    assignee_filter: str = Form("all"),
    q: str = Form(""),
):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    query = {"plan_date": {"$gte": week_start, "$lte": week_end}}
    meals_to_delete = await db.meal_plans.find(query, {"_id": 1}).to_list(length=None)
    meal_ids = [str(meal["_id"]) for meal in meals_to_delete]

    if meal_ids:
        await db.meal_checks.delete_many({"meal_plan_id": {"$in": meal_ids}})

    deleted = (await db.meal_plans.delete_many(query)).deleted_count
    context = await build_admin_meals_context(
        request,
        week_offset=week_offset,
        view_all=view_all,
        meal_type_filter=meal_type_filter,
        assignee_filter=assignee_filter,
        q=q,
        flash_message={"kind": "success", "text": f"Cleared {deleted} meals from the selected range."},
    )
    return templates.TemplateResponse("admin_meals_board.html", context)


@app.delete("/admin/meal/{meal_id}", response_class=HTMLResponse)
async def delete_meal(request: Request, meal_id: str):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    await db.meal_checks.delete_many({"meal_plan_id": meal_id})
    try:
        await db.meal_plans.delete_one({"_id": ObjectId(meal_id)})
    except Exception:
        pass # Invalid ID format, ignores block

    return HTMLResponse("")


@app.delete("/admin/meals/clear-week", response_class=HTMLResponse)
async def clear_week(request: Request, week_start: str = Query(...), week_end: str = Query(...)):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    
    query = {"plan_date": {"$gte": week_start, "$lte": week_end}}
    
    # Delete associated meal checks first
    meals_to_delete = await db.meal_plans.find(query, {"_id": 1}).to_list(length=None)
    meal_ids = [str(m["_id"]) for m in meals_to_delete]
    
    if meal_ids:
        await db.meal_checks.delete_many({"meal_plan_id": {"$in": meal_ids}})
        
    res = await db.meal_plans.delete_many(query)
    deleted = res.deleted_count

    return HTMLResponse(f'''
        <div class="upload-result success">
            <span class="icon">✅</span>
            <p>Cleared <strong>{deleted}</strong> meals from {week_start} to {week_end}</p>
            <p style="font-size:0.8rem; margin-top:0.5rem; color:var(--text-muted);">Refresh the page to see the updated table.</p>
        </div>
    ''')


@app.post("/admin/meal/add", response_class=HTMLResponse)
async def add_meal_manual(request: Request, plan_date: str = Form(...), meal_type: str = Form(...),
                          dish_name: str = Form(...), description: str = Form(""),
                          calories: int = Form(0), protein_g: float = Form(0),
                          carbs_g: float = Form(0), fat_g: float = Form(0), fiber_g: float = Form(0),
                          profile_id: str = Form("0")):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    pid = profile_id if profile_id and profile_id != "0" else None

    db = await get_db()
    
    res = await db.meal_plans.insert_one({
        "plan_date": plan_date,
        "meal_type": meal_type,
        "dish_name": dish_name.strip(),
        "description": description.strip(),
        "calories": calories,
        "protein_g": protein_g,
        "carbs_g": carbs_g,
        "fat_g": fat_g,
        "fiber_g": fiber_g,
        "profile_id": pid,
        "created_at": datetime.utcnow()
    })
    new_id = str(res.inserted_id)

    # Get all profiles for dropdown
    all_profiles = await db.meal_plans.database.profiles.find().sort("name", ASCENDING).to_list(length=None)
    
    # Build profile options
    profile_opts = '<option value="0"' + (' selected' if not pid else '') + '>👥 Everyone</option>'
    for p in all_profiles:
        p_id_str = str(p["_id"])
        sel = ' selected' if p_id_str == pid else ''
        profile_opts += f'<option value="{p_id_str}"{sel}>{p["name"]}</option>'

    return HTMLResponse(f"""
        <tr id="meal-row-{new_id}" class="new-row-flash">
            <td>
                <input type="date" value="{plan_date}" class="inline-input"
                       hx-post="/admin/meal/{new_id}/edit" hx-include="closest tr" hx-target="#meal-row-{new_id}" hx-swap="outerHTML"
                       name="plan_date">
            </td>
            <td>
                <select class="inline-input" name="meal_type"
                        hx-post="/admin/meal/{new_id}/edit" hx-include="closest tr" hx-target="#meal-row-{new_id}" hx-swap="outerHTML">
                    <option value="breakfast" {"selected" if meal_type == "breakfast" else ""}>🌅 Breakfast</option>
                    <option value="morning_snack" {"selected" if meal_type == "morning_snack" else ""}>🍎 Morning Snack</option>
                    <option value="lunch" {"selected" if meal_type == "lunch" else ""}>☀️ Lunch</option>
                    <option value="afternoon_snack" {"selected" if meal_type == "afternoon_snack" else ""}>🫐 Afternoon Snack</option>
                    <option value="dinner" {"selected" if meal_type == "dinner" else ""}>🌙 Dinner</option>
                    <option value="evening_snack" {"selected" if meal_type == "evening_snack" else ""}>🌜 Evening Snack</option>
                </select>
            </td>
            <td>
                <input type="text" value="{dish_name}" class="inline-input inline-name" name="dish_name"
                       hx-post="/admin/meal/{new_id}/edit" hx-include="closest tr" hx-target="#meal-row-{new_id}" hx-swap="outerHTML"
                       hx-trigger="change">
            </td>
            <td>
                <input type="number" value="{calories}" class="inline-input inline-num" name="calories"
                       hx-post="/admin/meal/{new_id}/edit" hx-include="closest tr" hx-target="#meal-row-{new_id}" hx-swap="outerHTML"
                       hx-trigger="change">
            </td>
            <td class="macros-cell">
                <input type="number" value="{protein_g}" class="inline-input inline-micro" name="protein_g" step="0.1" hx-post="/admin/meal/{new_id}/edit" hx-include="closest tr" hx-target="#meal-row-{new_id}" hx-swap="outerHTML" hx-trigger="change">
                <input type="number" value="{carbs_g}" class="inline-input inline-micro" name="carbs_g" step="0.1" hx-post="/admin/meal/{new_id}/edit" hx-include="closest tr" hx-target="#meal-row-{new_id}" hx-swap="outerHTML" hx-trigger="change">
                <input type="number" value="{fat_g}" class="inline-input inline-micro" name="fat_g" step="0.1" hx-post="/admin/meal/{new_id}/edit" hx-include="closest tr" hx-target="#meal-row-{new_id}" hx-swap="outerHTML" hx-trigger="change">
            </td>
            <td>
                <select class="inline-input inline-assign" name="profile_id"
                        hx-post="/admin/meal/{new_id}/edit" hx-include="closest tr" hx-target="#meal-row-{new_id}" hx-swap="outerHTML">
                    {profile_opts}
                </select>
            </td>
            <td class="actions-cell">
                <button class="btn btn-sm btn-copy" onclick="openCopyModal('{new_id}', '{dish_name}')">📋</button>
                <button class="btn btn-sm btn-danger"
                        hx-delete="/admin/meal/{new_id}"
                        hx-target="#meal-row-{new_id}"
                        hx-swap="outerHTML"
                        hx-confirm="Delete this meal?">🗑️</button>
            </td>
        </tr>
    """)


@app.post("/admin/meal/{meal_id}/edit", response_class=HTMLResponse)
async def edit_meal(request: Request, meal_id: str,
                    plan_date: str = Form(...), meal_type: str = Form(...),
                    dish_name: str = Form(...), calories: int = Form(0),
                    protein_g: float = Form(0), carbs_g: float = Form(0), fat_g: float = Form(0),
                    profile_id: str = Form("0")):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    pid = profile_id if profile_id and profile_id != "0" else None

    db = await get_db()
    try:
        await db.meal_plans.update_one(
            {"_id": ObjectId(meal_id)},
            {"$set": {
                "plan_date": plan_date,
                "meal_type": meal_type,
                "dish_name": dish_name.strip(),
                "calories": calories,
                "protein_g": protein_g,
                "carbs_g": carbs_g,
                "fat_g": fat_g,
                "profile_id": pid
            }}
        )
        meal_doc = await db.meal_plans.find_one({"_id": ObjectId(meal_id)})
    except Exception:
        return HTMLResponse("")

    if not meal_doc:
        return HTMLResponse("")

    # Get profiles for dropdown
    all_profiles = await db.profiles.find().sort("name", ASCENDING).to_list(length=None)

    mt = meal_doc["meal_type"]
    mpid = str(meal_doc.get("profile_id", "0"))

    profile_opts = '<option value="0"' + (' selected' if mpid == "0" or mpid == "None" else '') + '>👥 Everyone</option>'
    for p in all_profiles:
        p_id_str = str(p["_id"])
        sel = ' selected' if p_id_str == mpid else ''
        profile_opts += f'<option value="{p_id_str}"{sel}>{p["name"]}</option>'

    return HTMLResponse(f"""
        <tr id="meal-row-{meal_id}">
            <td>
                <input type="date" value="{meal_doc['plan_date']}" class="inline-input"
                       hx-post="/admin/meal/{meal_id}/edit" hx-include="closest tr" hx-target="#meal-row-{meal_id}" hx-swap="outerHTML"
                       name="plan_date">
            </td>
            <td>
                <select class="inline-input" name="meal_type"
                        hx-post="/admin/meal/{meal_id}/edit" hx-include="closest tr" hx-target="#meal-row-{meal_id}" hx-swap="outerHTML">
                    <option value="breakfast" {"selected" if mt == "breakfast" else ""}>🌅 Breakfast</option>
                    <option value="morning_snack" {"selected" if mt == "morning_snack" else ""}>🍎 Morning Snack</option>
                    <option value="lunch" {"selected" if mt == "lunch" else ""}>☀️ Lunch</option>
                    <option value="afternoon_snack" {"selected" if mt == "afternoon_snack" else ""}>🫐 Afternoon Snack</option>
                    <option value="dinner" {"selected" if mt == "dinner" else ""}>🌙 Dinner</option>
                    <option value="evening_snack" {"selected" if mt == "evening_snack" else ""}>🌜 Evening Snack</option>
                </select>
            </td>
            <td>
                <input type="text" value="{meal_doc['dish_name']}" class="inline-input inline-name" name="dish_name"
                       hx-post="/admin/meal/{meal_id}/edit" hx-include="closest tr" hx-target="#meal-row-{meal_id}" hx-swap="outerHTML"
                       hx-trigger="change">
            </td>
            <td>
                <input type="number" value="{meal_doc.get('calories', 0)}" class="inline-input inline-num" name="calories"
                       hx-post="/admin/meal/{meal_id}/edit" hx-include="closest tr" hx-target="#meal-row-{meal_id}" hx-swap="outerHTML"
                       hx-trigger="change">
            </td>
            <td class="macros-cell">
                <input type="number" value="{meal_doc.get('protein_g', 0)}" class="inline-input inline-micro" name="protein_g" step="0.1" hx-post="/admin/meal/{meal_id}/edit" hx-include="closest tr" hx-target="#meal-row-{meal_id}" hx-swap="outerHTML" hx-trigger="change">
                <input type="number" value="{meal_doc.get('carbs_g', 0)}" class="inline-input inline-micro" name="carbs_g" step="0.1" hx-post="/admin/meal/{meal_id}/edit" hx-include="closest tr" hx-target="#meal-row-{meal_id}" hx-swap="outerHTML" hx-trigger="change">
                <input type="number" value="{meal_doc.get('fat_g', 0)}" class="inline-input inline-micro" name="fat_g" step="0.1" hx-post="/admin/meal/{meal_id}/edit" hx-include="closest tr" hx-target="#meal-row-{meal_id}" hx-swap="outerHTML" hx-trigger="change">
            </td>
            <td>
                <select class="inline-input inline-assign" name="profile_id"
                        hx-post="/admin/meal/{meal_id}/edit" hx-include="closest tr" hx-target="#meal-row-{meal_id}" hx-swap="outerHTML">
                    {profile_opts}
                </select>
            </td>
            <td class="actions-cell">
                <button class="btn btn-sm btn-copy" onclick="openCopyModal('{meal_id}', '{meal_doc['dish_name']}')">📋</button>
                <button class="btn btn-sm btn-danger"
                        hx-delete="/admin/meal/{meal_id}"
                        hx-target="#meal-row-{meal_id}"
                        hx-swap="outerHTML"
                        hx-confirm="Delete this meal?">🗑️</button>
            </td>
        </tr>
    """)


@app.post("/admin/meal/{meal_id}/copy", response_class=HTMLResponse)
async def copy_meal(request: Request, meal_id: str, target_date: str = Form(...)):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    try:
        meal = await db.meal_plans.find_one({"_id": ObjectId(meal_id)})
        if not meal:
            return HTMLResponse('<div class="upload-result error"><span class="icon">❌</span><p>Meal not found.</p></div>')

        await db.meal_plans.insert_one({
            "plan_date": target_date,
            "meal_type": meal["meal_type"],
            "dish_name": meal["dish_name"],
            "description": meal.get("description", ""),
            "calories": meal.get("calories", 0),
            "protein_g": meal.get("protein_g", 0),
            "carbs_g": meal.get("carbs_g", 0),
            "fat_g": meal.get("fat_g", 0),
            "fiber_g": meal.get("fiber_g", 0),
            "profile_id": meal.get("profile_id"),
            "created_at": datetime.utcnow()
        })
    except Exception:
        return HTMLResponse('<div class="upload-result error"><span class="icon">❌</span><p>Error copying meal.</p></div>')

    return HTMLResponse(f'''
        <div class="upload-result success">
            <span class="icon">✅</span>
            <p>Copied "<strong>{meal["dish_name"]}</strong>" to <strong>{target_date}</strong></p>
        </div>
    ''')


@app.post("/admin/model/add", response_class=HTMLResponse)
async def add_model(request: Request, provider: str = Form(...), model_id: str = Form(...),
                    display_name: str = Form(...), api_key: str = Form(""),
                    search_grounding: str = Form("0"), include_youtube: str = Form("0")):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    await db.ai_models.insert_one({
        "provider": provider,
        "model_id": model_id.strip(),
        "display_name": display_name.strip(),
        "api_key": api_key.strip(),
        "is_default": 0,
        "search_grounding": 1 if provider == "gemini" and search_grounding == "1" else 0,
        "include_youtube": 1 if provider == "gemini" and include_youtube == "1" else 0,
        "created_at": datetime.utcnow()
    })

    return RedirectResponse(url="/admin#models", status_code=302)


@app.delete("/admin/model/{model_id_str}", response_class=HTMLResponse)
async def delete_model(request: Request, model_id_str: str):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    try:
        await db.ai_models.delete_one({"_id": ObjectId(model_id_str)})
    except Exception:
        pass

    return HTMLResponse("")


@app.post("/admin/model/default/{model_id_str}", response_class=HTMLResponse)
async def set_default_model(request: Request, model_id_str: str):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    try:
        await db.ai_models.update_many({}, {"$set": {"is_default": 0}})
        await db.ai_models.update_one({"_id": ObjectId(model_id_str)}, {"$set": {"is_default": 1}})
    except Exception:
        pass

    return RedirectResponse(url="/admin#models", status_code=302)


@app.get("/admin/activity", response_class=HTMLResponse)
async def admin_activity(request: Request, profile_filter: str = "0", page: int = 1):
    if not get_admin_session(request):
        return RedirectResponse(url="/admin/login", status_code=302)

    per_page = 100
    offset = (page - 1) * per_page

    db = await get_db()
    
    # Needs a manual lookup join equivalent since MongoDB doesn't easily JOIN outside of aggregations
    all_profiles = await db.profiles.find().sort("name", ASCENDING).to_list(length=None)
    profile_map = {str(p["_id"]): p["name"] for p in all_profiles}

    query = {}
    if profile_filter and profile_filter != "0":
        query["profile_id"] = profile_filter

    recent_acts = await db.activity_logs.find(query).sort("created_at", DESCENDING).skip(offset).limit(per_page).to_list(None)
    
    # Attach profile_name manually
    activities = []
    for act in recent_acts:
        act["profile_name"] = profile_map.get(str(act.get("profile_id")), "Unknown Profile")
        activities.append(act)
        
    for p in all_profiles:
        p["id"] = str(p["_id"])

    return templates.TemplateResponse("admin_activity.html", {
        "request": request,
        "activities": activities,
        "profiles": all_profiles,
        "profile_filter": profile_filter,
        "page": page,
    })


@app.post("/admin/password", response_class=HTMLResponse)
async def change_admin_password(request: Request, current_password: str = Form(...), new_password: str = Form(...)):
    if not get_admin_session(request):
        return HTMLResponse("Unauthorized", status_code=401)

    db = await get_db()
    admin = await db.admin.find_one({})
    
    if not bcrypt.verify(current_password, admin["password_hash"]):
        return HTMLResponse("""<div class="upload-result error"><span class="icon">❌</span><p>Current password incorrect.</p></div>""")
    
    new_hash = bcrypt.hash(new_password)
    await db.admin.update_one({"_id": admin["_id"]}, {"$set": {"password_hash": new_hash}})

    return HTMLResponse("""<div class="upload-result success"><span class="icon">✅</span><p>Password updated successfully.</p></div>""")
